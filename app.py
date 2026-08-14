import os
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="Ratio Confirmation",
    page_icon="📊",
    layout="centered"
)


# =========================================================
# CREATE RATIO LIST
# =========================================================

def create_ratio_list(ratio_input):

    ratio_input = ratio_input.strip()

    try:

        if "-" in ratio_input:

            parts = ratio_input.split("-")

            if len(parts) != 2:
                raise ValueError

            start = int(parts[0])
            end = int(parts[1])

            if start > end:
                raise ValueError

            ratio_list = list(
                range(start, end + 1, 5)
            )

        else:

            ratio_list = [int(ratio_input)]

    except Exception:

        raise Exception(
            "Invalid Ratio. Please enter 75 or 20-60."
        )

    if any(r <= 0 for r in ratio_list):

        raise Exception(
            "Ratio must be greater than 0."
        )

    return ratio_list


# =========================================================
# PROCESS EXCEL
# =========================================================

def process_excel(
    uploaded_file,
    service,
    ratio_input,
    progress_bar,
    status
):

    # -----------------------------------------------------
    # RATIO
    # -----------------------------------------------------

    ratio_list = create_ratio_list(
        ratio_input
    )

    status.write(
        f"🔢 Ratio List: {ratio_list}"
    )

    progress_bar.progress(5)


    # -----------------------------------------------------
    # LOAD EXCEL
    # -----------------------------------------------------

    status.write(
        "📂 Loading Excel file..."
    )

    input_name = uploaded_file.name

    file_data = uploaded_file.getvalue()

    keep_vba = input_name.lower().endswith(".xlsm")

    wb = load_workbook(
        BytesIO(file_data),
        keep_vba=keep_vba
    )

    if not wb.sheetnames:

        raise Exception(
            "Excel file does not contain any sheet."
        )

    base_ws = wb[wb.sheetnames[0]]

    progress_bar.progress(10)


    # -----------------------------------------------------
    # CREATE RATIO SHEETS
    # -----------------------------------------------------

    status.write(
        "📄 Creating ratio sheets..."
    )

    base_ws.title = str(
        ratio_list[0]
    )

    for r in ratio_list[1:]:

        new_ws = wb.copy_worksheet(
            base_ws
        )

        new_ws.title = str(r)

    progress_bar.progress(15)


    # -----------------------------------------------------
    # MONTH NAMES
    # -----------------------------------------------------

    months = [
        "Jan'", "Feb'", "Mar'", "Apr'",
        "May'", "Jun'", "Jul'", "Aug'",
        "Sep'", "Oct'", "Nov'", "Dec'"
    ]


    # -----------------------------------------------------
    # PROCESS EACH RATIO SHEET
    # -----------------------------------------------------

    total_ratios = len(ratio_list)

    for index, ratio in enumerate(
        ratio_list
    ):

        progress = 15 + int(
            (index / total_ratios) * 65
        )

        progress_bar.progress(
            progress
        )

        status.write(
            f"⚙️ Processing Ratio: {ratio}"
        )

        div = ratio + 1

        ws = wb[str(ratio)]


        # -------------------------------------------------
        # FIND MONTH COLUMNS
        # -------------------------------------------------

        mcols = []

        for c in range(
            1,
            ws.max_column + 1
        ):

            h = str(
                ws.cell(1, c).value or ""
            )

            if any(
                m in h
                for m in months
            ):

                mcols.append(c)

        if not mcols:

            status.write(
                f"⚠️ No month columns found in Ratio {ratio}"
            )

            continue


        # -------------------------------------------------
        # TOTAL CANDIDATE & MAX CANDIDATE
        # -------------------------------------------------

        tot = ws.max_column + 1

        mx = tot + 1

        ws.cell(
            1,
            tot
        ).value = "Total Candidate"

        ws.cell(
            1,
            mx
        ).value = "Max Candidate"


        for r in range(
            2,
            ws.max_row + 1
        ):

            refs = [
                f"{get_column_letter(c)}{r}"
                for c in mcols
            ]

            ws.cell(
                r,
                tot
            ).value = (
                f"=SUM({','.join(refs)})"
            )

            ws.cell(
                r,
                mx
            ).value = (
                f"=MAX({','.join(refs)})"
            )


        # -------------------------------------------------
        # OPR COLUMNS
        # -------------------------------------------------

        ins = mx + 1

        opr = []

        for i, oc in enumerate(
            mcols
        ):

            ws.insert_cols(
                ins + i
            )

            nc = ins + i

            hdr = ws.cell(
                1,
                oc
            ).value

            ws.cell(
                1,
                nc
            ).value = f"{hdr} Opr"

            opr.append(nc)

            col = get_column_letter(
                oc
            )

            for r in range(
                2,
                ws.max_row + 1
            ):

                ws.cell(
                    r,
                    nc
                ).value = (
                    f"=ROUNDUP("
                    f"{col}{r}/{div},0)"
                )


        # -------------------------------------------------
        # MAX OPR
        # -------------------------------------------------

        maxopr = ws.max_column + 1

        ws.cell(
            1,
            maxopr
        ).value = "Max Opr"


        for r in range(
            2,
            ws.max_row + 1
        ):

            refs = [
                f"{get_column_letter(c)}{r}"
                for c in opr
            ]

            ws.cell(
                r,
                maxopr
            ).value = (
                f"=MAX({','.join(refs)})"
            )


        # -------------------------------------------------
        # SERVICE COLUMNS
        # -------------------------------------------------

        if service.upper() == "FPS":

            names = [
                "Tab",
                "FPS",
                "OTG",
                "Hologram",
                "Id Card",
                "Jacket"
            ]

        elif service.upper() == "IRIS":

            names = [
                "Tab",
                "IRIS",
                "FPS",
                "OTG",
                "Hologram",
                "Id Card",
                "Jacket"
            ]

        else:

            raise Exception(
                "Service must be FPS or IRIS."
            )


        start = ws.max_column + 1

        cols = {}

        for i, name in enumerate(
            names
        ):

            col = start + i

            cols[name] = col

            ws.cell(
                1,
                col
            ).value = name


        maxopr_letter = get_column_letter(
            maxopr
        )

        total_letter = get_column_letter(
            tot
        )

        tab_letter = get_column_letter(
            cols["Tab"]
        )


        # -------------------------------------------------
        # SERVICE FORMULAS
        # -------------------------------------------------

        for r in range(
            2,
            ws.max_row + 1
        ):

            # TAB

            ws.cell(
                r,
                cols["Tab"]
            ).value = (
                f"=IF({maxopr_letter}{r}=0,0,"
                f"IF(AND("
                f"{maxopr_letter}{r}>=8,"
                f"{maxopr_letter}{r}<16),"
                f"{maxopr_letter}{r}+2,"
                f"IF("
                f"{maxopr_letter}{r}>15,"
                f"{maxopr_letter}{r}+3,"
                f"{maxopr_letter}{r}+1"
                f")))"
            )


            # FPS / IRIS

            if service.upper() == "FPS":

                ws.cell(
                    r,
                    cols["FPS"]
                ).value = (
                    f"={tab_letter}{r}"
                )

            else:

                ws.cell(
                    r,
                    cols["IRIS"]
                ).value = (
                    f"={tab_letter}{r}"
                )

                ws.cell(
                    r,
                    cols["FPS"]
                ).value = "=1"


            # OTG

            ws.cell(
                r,
                cols["OTG"]
            ).value = (
                f"={tab_letter}{r}*2"
            )


            # HOLOGRAM

            ws.cell(
                r,
                cols["Hologram"]
            ).value = (
                f"=ROUNDUP("
                f"{total_letter}{r}/100,0)+1"
            )


            # ID CARD

            ws.cell(
                r,
                cols["Id Card"]
            ).value = (
                f"={maxopr_letter}{r}+1"
            )


            # JACKET

            ws.cell(
                r,
                cols["Jacket"]
            ).value = (
                f"={maxopr_letter}{r}"
            )


    # =====================================================
    # CREATE SUMMARY
    # =====================================================

    status.write(
        "📊 Creating Summary sheet..."
    )

    progress_bar.progress(82)


    if "Summary" in wb.sheetnames:

        del wb["Summary"]


    summary = wb.create_sheet(
        "Summary",
        0
    )


    first_ws = wb[
        str(ratio_list[0])
    ]


    # -----------------------------------------------------
    # FIND MONTH OPR COLUMNS
    # -----------------------------------------------------

    month_headers = []

    opr_columns = []

    for c in range(
        1,
        first_ws.max_column + 1
    ):

        h = str(
            first_ws.cell(1, c).value or ""
        ).strip()

        if (
            h.endswith(" Opr")
            and h != "Max Opr"
        ):

            month_headers.append(h)

            opr_columns.append(c)


    # -----------------------------------------------------
    # HEADER MAP
    # -----------------------------------------------------

    header_map = {}

    for c in range(
        1,
        first_ws.max_column + 1
    ):

        header_map[
            str(
                first_ws.cell(
                    1,
                    c
                ).value
            )
        ] = c


    if "Total Candidate" not in header_map:

        raise Exception(
            "Total Candidate column not found."
        )

    if "Max Candidate" not in header_map:

        raise Exception(
            "Max Candidate column not found."
        )

    if "Max Opr" not in header_map:

        raise Exception(
            "Max Opr column not found."
        )


    tot_col = header_map[
        "Total Candidate"
    ]

    maxcand_col = header_map[
        "Max Candidate"
    ]

    maxopr_col = header_map[
        "Max Opr"
    ]


    # -----------------------------------------------------
    # SERVICE COLUMN MAP
    # -----------------------------------------------------

    service_cols = {}

    for name in [
        "Tab",
        "FPS",
        "IRIS",
        "OTG",
        "Hologram",
        "Id Card",
        "Jacket"
    ]:

        if name in header_map:

            service_cols[name] = (
                header_map[name]
            )


    # -----------------------------------------------------
    # SUMMARY HEADERS
    # -----------------------------------------------------

    headers = [
        "Ratio",
        "Total Center",
        "Total Candidate",
        "Max Candidate"
    ]

    headers += month_headers

    headers += [
        "Max Opr"
    ]


    if service.upper() == "FPS":

        headers += [
            "Tab",
            "FPS",
            "OTG",
            "Hologram",
            "Id Card",
            "Jacket"
        ]

    else:

        headers += [
            "Tab",
            "IRIS",
            "FPS",
            "OTG",
            "Hologram",
            "Id Card",
            "Jacket"
        ]

    headers.append(
        "Avg"
    )


    for c, h in enumerate(
        headers,
        1
    ):

        summary.cell(
            1,
            c
        ).value = h


    # -----------------------------------------------------
    # SUMMARY DATA
    # -----------------------------------------------------

    for row, ratio in enumerate(
        ratio_list,
        2
    ):

        sh = f"'{ratio}'"

        col = 1


        # Ratio

        summary.cell(
            row,
            col
        ).value = ratio

        col += 1


        # Total Center

        letter = get_column_letter(
            tot_col
        )

        summary.cell(
            row,
            col
        ).value = (
            f"=COUNT("
            f"{sh}!{letter}:{letter}"
            f")"
        )

        col += 1


        # Total Candidate

        summary.cell(
            row,
            col
        ).value = (
            f"=SUM("
            f"{sh}!{letter}:{letter}"
            f")"
        )

        col += 1


        # Max Candidate

        letter = get_column_letter(
            maxcand_col
        )

        summary.cell(
            row,
            col
        ).value = (
            f"=SUM("
            f"{sh}!{letter}:{letter}"
            f")"
        )

        col += 1


        # Monthly OPR

        for oc in opr_columns:

            letter = get_column_letter(
                oc
            )

            summary.cell(
                row,
                col
            ).value = (
                f"=SUM("
                f"{sh}!{letter}:{letter}"
                f")"
            )

            col += 1


        # Max OPR

        letter = get_column_letter(
            maxopr_col
        )

        summary.cell(
            row,
            col
        ).value = (
            f"=SUM("
            f"{sh}!{letter}:{letter}"
            f")"
        )

        col += 1


        # Service order

        order = [
            "Tab"
        ]

        if service.upper() == "IRIS":

            order.append(
                "IRIS"
            )

        order += [
            "FPS",
            "OTG",
            "Hologram",
            "Id Card",
            "Jacket"
        ]


        for name in order:

            if name not in service_cols:

                continue

            letter = get_column_letter(
                service_cols[name]
            )

            summary.cell(
                row,
                col
            ).value = (
                f"=SUM("
                f"{sh}!{letter}:{letter}"
                f")"
            )

            col += 1


        # Average

        last_data_col = col - 1

        if last_data_col >= 2:

            summary.cell(
                row,
                col
            ).value = (
                f"=AVERAGE("
                f"B{row}:"
                f"{get_column_letter(last_data_col)}{row}"
                f")"
            )


    # -----------------------------------------------------
    # FORMAT SUMMARY
    # -----------------------------------------------------

    for cell in summary[1]:

        cell.font = cell.font.copy(
            bold=True
        )


    summary.freeze_panes = "A2"


    # -----------------------------------------------------
    # EXCEL CALCULATION SETTINGS
    # -----------------------------------------------------

    try:

        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"

    except Exception:

        pass


    # =====================================================
    # SAVE OUTPUT
    # =====================================================

    status.write(
        "💾 Preparing output Excel..."
    )

    progress_bar.progress(95)


    output = BytesIO()

    wb.save(output)

    wb.close()

    output.seek(0)


    # -----------------------------------------------------
    # OUTPUT NAME
    # -----------------------------------------------------

    base_name = os.path.splitext(
        input_name
    )[0]

    output_name = (
        f"{base_name}_Output.xlsx"
    )


    progress_bar.progress(100)

    status.write(
        "✅ Processing completed successfully!"
    )


    return output, output_name


# =========================================================
# STREAMLIT USER INTERFACE
# =========================================================

st.title(
    "📊 Ratio Confirmation"
)

st.write(
    "Upload your Excel file and generate the Ratio Confirmation output."
)


# =========================================================
# 1. FILE UPLOAD
# =========================================================

st.subheader(
    "1. 📁 Excel Upload"
)

uploaded_file = st.file_uploader(
    "Drag & Drop Excel file here",
    type=[
        "xlsx",
        "xlsm"
    ],
    help="Upload an Excel .xlsx or .xlsm file."
)


# =========================================================
# SHOW FILE NAME
# =========================================================

if uploaded_file is not None:

    st.success(
        f"📄 Selected file: {uploaded_file.name}"
    )


# =========================================================
# 2. SERVICE
# =========================================================

st.subheader(
    "2. 📋 Service"
)

service = st.selectbox(
    "Select Service",
    [
        "FPS",
        "IRIS"
    ]
)


# =========================================================
# 3. RATIO
# =========================================================

st.subheader(
    "3. 🔢 Ratio"
)

ratio_input = st.text_input(
    "Enter Ratio",
    placeholder="Example: 75 or 20-60"
)


# =========================================================
# PROCESS BUTTON
# =========================================================

st.subheader(
    "4. ⚙️ Processing"
)

process_button = st.button(
    "🚀 Process Excel",
    type="primary",
    use_container_width=True
)


# =========================================================
# PROCESS
# =========================================================

if process_button:

    if uploaded_file is None:

        st.error(
            "❌ Please upload an Excel file first."
        )

    elif not ratio_input.strip():

        st.error(
            "❌ Please enter a ratio. Example: 75 or 20-60."
        )

    else:

        progress_bar = st.progress(
            0
        )

        status = st.empty()

        try:

            output_file, output_name = process_excel(
                uploaded_file,
                service,
                ratio_input,
                progress_bar,
                status
            )

            st.success(
                "🎉 Excel processing completed successfully!"
            )

            st.download_button(
                label="📥 Download Excel",
                data=output_file,
                file_name=output_name,
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
                use_container_width=True
            )

        except Exception as e:

            st.error(
                f"❌ Processing Error: {str(e)}"
            )
